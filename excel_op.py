import openpyxl

wb = openpyxl.load_workbook('example.xlsx')
sheet = wb.get_active_sheet()
print('此sheet有%s行,%s列数据' % (sheet.max_row, sheet.max_column))
# for row_cell_objects in sheet['A1':'C1']:
#     for cell_object in row_cell_objects:
#         # cell有value，row、column 和 coordinate 属性
#         print(cell_object.coordinate, cell_object.value)
#         print('--- END OF ROW ---')
for i in range(1, sheet.max_row):
    for j in range(1, sheet.max_column):
        print(sheet.cell(row=i, column=j).value)
    print('--- END OF ROW ---')