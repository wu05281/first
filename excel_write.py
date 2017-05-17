import openpyxl

wb = openpyxl.Workbook()
#sheet = wb.create_sheet(0)
sheet1 = wb.get_active_sheet()
sheet1['A1'] = 'hello world'
sheet1.title = 'hello'
sheet2 = wb.create_sheet('haha')
middle_sheet =wb.create_sheet(index=1, title='Middle Sheet')
wb.remove_sheet(wb.get_sheet_by_name('Middle Sheet'))
wb.save('example_copy.xlsx')