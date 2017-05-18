import openpyxl
from openpyxl.styles import Font

wb = openpyxl.load_workbook('example.xlsx')
sheet = wb.get_sheet_by_name('工作表1')
#设置excel样式
italic24_font = Font(size=24, italic=True)
col = sheet.column_dimensions['A']
col.width = 20

PRICE_UPDATES = {'Garlic': 3.07, 'Celery': 1.19, 'Lemon': 1.27}
for i in range(2, sheet.max_row+1):
    product_name = sheet.cell(row=i, column=1).value
    if product_name in PRICE_UPDATES:
        sheet.cell(row=i, column=2).value = PRICE_UPDATES[product_name]
wb.save('example_update.xlsx')